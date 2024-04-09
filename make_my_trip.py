from selenium import webdriver      
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from openpyxl import Workbook, load_workbook
import time , sys
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, StaleElementReferenceException


def is_element_exists(x_path):
    try:
        driver.find_element(By.XPATH,x_path)
        # print(x_path,'  True')
        return True
    except NoSuchElementException:
        # print(x_path,'  NoSuchElementException')
        return False
    
wb = Workbook()
ws = wb.active

options = webdriver.ChromeOptions()
driver = webdriver.Chrome(options=options)
driver.get('https://www.makemytrip.com/bus-tickets/city-index_a_1.html')
time.sleep(5)

page_no_elements_list = driver.find_elements(By.XPATH, "/html/body/div/div[2]/div/div[1]/main/div[2]/div")
time.sleep(2)

no_pages_to_fetch_data = driver.find_elements(By.XPATH, '/html/body/div/div[2]/div/div[1]/main/div[1]/div[2]/a')
time.sleep(5)
print(len(no_pages_to_fetch_data))

for k in range(1, len(no_pages_to_fetch_data)+1):

    driver.find_element(By.XPATH, '/html/body/div/div[2]/div/div[1]/main/div[1]/div[2]/a['+str(k)+']').click()
    time.sleep(5)
    
    heading_of_each_sheets = driver.find_element(By.XPATH, '/html/body/div/div[2]/div/div[1]/main/div[1]/div[2]/a['+str(k)+']').text
    ws = wb.create_sheet(heading_of_each_sheets)
    
    driver.find_element(By.XPATH, "/html/body/div/div[2]/div/div[1]/div[1]/div").click()
    time.sleep(2)
    for i in range(len(page_no_elements_list)):
        elements = driver.find_elements(By.XPATH, '/html/body/div/div[2]/div/div[1]/main/div[1]/section/div/div')
        print(len(elements))
         
        for j in elements:
           ws.append([j.text])
            
        if is_element_exists("//a[text()  = 'next ›']"):
            driver.find_element(By.XPATH, "//a[text()  = 'next ›']").click()  
        else:
            break
wb.save('trip.xlsx')


    
# def wait_and_click(x_path):
#     wait_till_element_exist(x_path)
#     i = 1
#     while True:
#         try:
#             driver.find_element(By.XPATH,x_path).click()
#             return True
#             # driver.find_element(By.XPATH,x_path).send_keys(Keys.ENTER)
#         except StaleElementReferenceException:
#             print(x_path,'  StaleElementReferenceException')
#             time.sleep(2)
#         except ElementClickInterceptedException:
#             print(x_path,'  ElementClickInterceptedException')
#             time.sleep(2)
#         except ElementNotInteractableException:
#             print(x_path,'  ElementClickInterceptedException')
#             time.sleep(2)
#         if i >= config.wait_time:
#             return False
#         i += 1
        
# def wait_till_element_exist(x_path):
#     i = 0
#     while True:
#         if is_element_exists(x_path):
#             return True
#         else:
#             time.sleep(2)
#             if i >= config.wait_time:
#                 print('Element ',x_path,' does not exists')
#                 return False
#         i += 1
        
# def wait_and_enter(x_path,key):
#     wait_till_element_exist(x_path)
#     i = 1
#     while True:
#         try:
#             driver.find_element(By.XPATH,x_path).send_keys(key)
#             return True
#         except StaleElementReferenceException:
#             print(x_path, 'StaleElementReferenceException')
#             time.sleep(2)
#         except ElementNotInteractableException:
#             print(x_path, 'ElementNotInteractableException')
#             time.sleep(2)
#         if i > config.wait_time:
#             return False
#         i += 1
        
        
        
        
    